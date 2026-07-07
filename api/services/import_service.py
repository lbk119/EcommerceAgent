"""数据导入服务。"""

from __future__ import annotations

import csv
import hashlib
import json
import uuid
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

from fastapi import UploadFile

from api.db import execute, execute_many, fetch_one
from api.services.ecommerce_queries import list_import_jobs
from api.services.post_import_service import run_post_import_refresh


IMPORT_DIR = Path(__file__).resolve().parents[2] / "updated" / "data_imports"
IMPORT_DIR.mkdir(parents=True, exist_ok=True)


def create_sample_job(tenant_id: str, shop_id: str, user_id: str) -> dict[str, Any]:
    """给当前店铺写入一批可测试的模拟经营数据。"""
    job_id = str(uuid.uuid4())
    seeded_rows = seed_sample_operating_data(tenant_id, shop_id)
    execute(
        """
        INSERT INTO data_import_jobs (id, tenant_id, shop_id, source, file_name, status, rows_count, quality_score, mapping_json, created_by)
        VALUES (%s, %s, %s, 'sample', 'orange_shop_sample_import.csv', 'imported', %s, 96, %s, %s)
        """,
        (job_id, tenant_id, shop_id, seeded_rows, json.dumps({"mode": "sample", "description": "内置模拟订单、商品、库存、流量、活动和退款数据"}, ensure_ascii=False), user_id),
    )
    refresh_result = run_post_import_refresh(tenant_id, shop_id, user_id, job_id)
    return {"id": job_id, "status": "imported", "rows": seeded_rows, **refresh_result}


def seed_sample_operating_data(tenant_id: str, shop_id: str) -> int:
    """把内置样例写入当前租户/店铺的经营表。

    这里使用稳定 ID 和 ON DUPLICATE KEY UPDATE，方便你反复点击“使用示例数据”测试，而不会
    因主键重复导致失败。真实 CSV 入库后续可以复用同样的表写入策略。
    """
    prefix = _scope_prefix(tenant_id, shop_id)
    now = datetime.now().replace(microsecond=0)
    products = [
        {"id": f"{prefix}_orange_001", "name": "赣南脐橙礼盒 5kg", "category": "fresh_orange", "price": 89.0, "stock": 48, "safety": 120, "visitors": 3200, "conversions": 286},
        {"id": f"{prefix}_orange_002", "name": "冰糖橙家庭装 9斤", "category": "fresh_orange", "price": 59.0, "stock": 860, "safety": 180, "visitors": 2480, "conversions": 164},
        {"id": f"{prefix}_juice_003", "name": "鲜榨橙汁 NFC 12瓶", "category": "juice", "price": 129.0, "stock": 210, "safety": 90, "visitors": 1820, "conversions": 92},
        {"id": f"{prefix}_gift_004", "name": "企业福利橙礼盒", "category": "gift_box", "price": 168.0, "stock": 26, "safety": 80, "visitors": 960, "conversions": 61},
        {"id": f"{prefix}_orange_005", "name": "小果试吃装 3斤", "category": "fresh_orange", "price": 29.9, "stock": 1320, "safety": 150, "visitors": 4100, "conversions": 188},
        {"id": f"{prefix}_jam_006", "name": "手工橙皮果酱 2瓶", "category": "orange_jam", "price": 49.0, "stock": 340, "safety": 100, "visitors": 760, "conversions": 24},
    ]

    execute_many(
        """
        INSERT INTO customers (tenant_id, shop_id, customer_id, customer_unique_id, customer_zip_code_prefix, customer_city, customer_state)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE customer_city=VALUES(customer_city), customer_state=VALUES(customer_state)
        """,
        [(tenant_id, shop_id, f"{prefix}_cust_{index:03d}", f"{prefix}_unique_{index:03d}", 330000 + index, city, state) for index, (city, state) in enumerate([
            ("杭州", "浙江"), ("上海", "上海"), ("广州", "广东"), ("深圳", "广东"), ("成都", "四川"), ("武汉", "湖北"), ("南京", "江苏"), ("北京", "北京"),
        ], start=1)],
    )
    execute(
        """
        INSERT INTO sellers (tenant_id, shop_id, seller_id, seller_zip_code_prefix, seller_city, seller_state)
        VALUES (%s, %s, %s, 341000, '赣州', '江西')
        ON DUPLICATE KEY UPDATE seller_city=VALUES(seller_city), seller_state=VALUES(seller_state)
        """,
        (tenant_id, shop_id, f"{prefix}_seller_001"),
    )
    execute_many(
        """
        INSERT INTO products (tenant_id, shop_id, product_id, category_name, category_name_en, product_name_length, product_description_length, product_photos_qty, product_weight_g, product_length_cm, product_height_cm, product_width_cm)
        VALUES (%s, %s, %s, %s, %s, %s, 260, 5, 5000, 35, 22, 28)
        ON DUPLICATE KEY UPDATE category_name=VALUES(category_name), category_name_en=VALUES(category_name_en)
        """,
        [(tenant_id, shop_id, product["id"], product["name"], product["category"], len(product["name"])) for product in products],
    )
    execute_many(
        """
        INSERT INTO inventory (tenant_id, shop_id, product_id, stock, safety_stock, warehouse, updated_at)
        VALUES (%s, %s, %s, %s, %s, '华东前置仓', %s)
        ON DUPLICATE KEY UPDATE stock=VALUES(stock), safety_stock=VALUES(safety_stock), updated_at=VALUES(updated_at)
        """,
        [(tenant_id, shop_id, product["id"], product["stock"], product["safety"], now) for product in products],
    )
    execute_many(
        """
        INSERT INTO traffic_stats (tenant_id, shop_id, stat_date, product_id, views, visitors, add_to_cart, favorites, conversions)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """,
        [(tenant_id, shop_id, (now - timedelta(days=day)).date(), product["id"], product["visitors"] * 2, product["visitors"], max(12, product["conversions"] * 2), max(8, product["conversions"]), product["conversions"]) for day in range(1, 4) for product in products],
    )

    order_rows = []
    order_item_rows = []
    payment_rows = []
    refund_rows = []
    for index in range(1, 37):
        product = products[index % len(products)]
        quantity = 1 + (index % 3 == 0)
        order_id = f"{prefix}_order_{index:04d}"
        purchase_time = now - timedelta(hours=index * 3)
        order_rows.append((tenant_id, shop_id, order_id, f"{prefix}_cust_{(index % 8) + 1:03d}", "delivered", purchase_time, purchase_time + timedelta(minutes=8), purchase_time + timedelta(hours=16), purchase_time + timedelta(days=2), purchase_time + timedelta(days=4)))
        order_item_rows.append((tenant_id, shop_id, order_id, 1, product["id"], f"{prefix}_seller_001", purchase_time + timedelta(days=1), product["price"] * quantity, 8.0 + (index % 4)))
        payment_rows.append((tenant_id, shop_id, order_id, 1, "credit_card" if index % 2 else "wallet", 1, product["price"] * quantity + 8.0 + (index % 4)))
        if index in {7, 18, 29}:
            refund_rows.append((tenant_id, shop_id, f"{prefix}_refund_{index:04d}", order_id, product["id"], purchase_time + timedelta(days=3), round(product["price"] * 0.35, 2), "果品破损/坏果", "approved"))

    execute_many(
        """
        INSERT INTO orders (tenant_id, shop_id, order_id, customer_id, order_status, order_purchase_timestamp, order_approved_at, order_delivered_carrier_date, order_delivered_customer_date, order_estimated_delivery_date)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE order_status=VALUES(order_status), order_purchase_timestamp=VALUES(order_purchase_timestamp)
        """,
        order_rows,
    )
    execute_many(
        """
        INSERT INTO order_items (tenant_id, shop_id, order_id, order_item_id, product_id, seller_id, shipping_limit_date, price, freight_value)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE product_id=VALUES(product_id), price=VALUES(price), freight_value=VALUES(freight_value)
        """,
        order_item_rows,
    )
    execute_many(
        """
        INSERT INTO payments (tenant_id, shop_id, order_id, payment_sequential, payment_type, payment_installments, payment_value)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE payment_type=VALUES(payment_type), payment_value=VALUES(payment_value)
        """,
        payment_rows,
    )
    execute_many(
        """
        INSERT INTO refunds (tenant_id, shop_id, refund_id, order_id, product_id, refund_time, refund_amount, refund_reason, refund_status)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE refund_status=VALUES(refund_status), refund_amount=VALUES(refund_amount)
        """,
        refund_rows,
    )

    campaign_618_id = _short_scoped_id(prefix, "campaign", "618 橘子礼盒直播专场", max_len=32)
    campaign_member_id = _short_scoped_id(prefix, "campaign", "会员日鲜橙满减", max_len=32)
    campaigns = [
        (tenant_id, shop_id, campaign_618_id, "618 橘子礼盒直播专场", "抖音直播", now - timedelta(days=10), now + timedelta(days=2), 18000.0, "active"),
        (tenant_id, shop_id, campaign_member_id, "会员日鲜橙满减", "淘宝 / 天猫", now - timedelta(days=6), now + timedelta(days=1), 8000.0, "active"),
    ]
    execute_many(
        """
        INSERT INTO campaigns (tenant_id, shop_id, campaign_id, campaign_name, channel, start_time, end_time, budget, status)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE campaign_name=VALUES(campaign_name), status=VALUES(status)
        """,
        campaigns,
    )
    execute_many(
        """
        INSERT INTO campaign_product_stats (tenant_id, shop_id, campaign_id, product_id, impressions, clicks, orders_count, revenue, spend)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE impressions=VALUES(impressions), clicks=VALUES(clicks), orders_count=VALUES(orders_count), revenue=VALUES(revenue), spend=VALUES(spend)
        """,
        [(tenant_id, shop_id, campaigns[index % len(campaigns)][2], product["id"], product["visitors"] * 3, product["visitors"], product["conversions"], product["price"] * product["conversions"], max(300.0, product["price"] * product["conversions"] / 4)) for index, product in enumerate(products)],
    )
    return len(order_rows) + len(products) + len(campaigns) + len(refund_rows)


async def create_upload_job(tenant_id: str, shop_id: str, user_id: str, file: UploadFile) -> dict[str, Any]:
    """保存上传文件并创建导入 job。解析和入库通过 preview/mapping/confirm 分阶段完成。"""
    job_id = str(uuid.uuid4())
    safe_name = Path(file.filename or "upload.csv").name
    target = IMPORT_DIR / f"{job_id}_{safe_name}"
    with target.open("wb") as buffer:
        buffer.write(await file.read())
    execute(
        """
        INSERT INTO data_import_jobs (id, tenant_id, shop_id, source, file_name, status, rows_count, quality_score, created_by)
        VALUES (%s, %s, %s, 'upload', %s, 'mapping_required', 0, 0, %s)
        """,
        (job_id, tenant_id, shop_id, safe_name, user_id),
    )
    return {"id": job_id, "fileName": safe_name, "status": "mapping_required"}


def create_paste_job(tenant_id: str, shop_id: str, user_id: str, text: str) -> dict[str, Any]:
    """保存粘贴的 CSV/TSV 文本并创建导入 job，后续复用 preview/confirm。"""
    cleaned = (text or "").strip()
    if not cleaned:
        raise ValueError("粘贴内容为空")
    job_id = str(uuid.uuid4())
    safe_name = "paste_import.tsv" if "\t" in cleaned.splitlines()[0] else "paste_import.csv"
    target = IMPORT_DIR / f"{job_id}_{safe_name}"
    target.write_text(cleaned + "\n", encoding="utf-8-sig")
    preview_rows = _read_upload_rows(target, limit=20)
    execute(
        """
        INSERT INTO data_import_jobs (id, tenant_id, shop_id, source, file_name, status, rows_count, quality_score, created_by)
        VALUES (%s, %s, %s, 'paste', %s, 'mapping_required', %s, %s, %s)
        """,
        (job_id, tenant_id, shop_id, safe_name, len(preview_rows), 88 if preview_rows else 0, user_id),
    )
    return {"id": job_id, "fileName": safe_name, "status": "mapping_required"}


def preview_job(tenant_id: str, shop_id: str, job_id: str) -> dict[str, Any]:
    """返回上传文件前 20 行和简单字段识别结果。"""
    row = fetch_one("SELECT file_name FROM data_import_jobs WHERE tenant_id=%s AND shop_id=%s AND id=%s", (tenant_id, shop_id, job_id))
    if not row:
        return {"rows": [], "fields": [], "quality": {"score": 0, "errors": ["导入任务不存在"]}}
    matches = list(IMPORT_DIR.glob(f"{job_id}_*"))
    if not matches:
        return {"rows": [], "fields": [], "quality": {"score": 0, "errors": ["上传文件不存在"]}}
    preview_rows = _read_upload_rows(matches[0], limit=20)
    fields = [{"sourceField": field, "targetField": _guess_target(field), "confidence": 0.82} for field in (preview_rows[0].keys() if preview_rows else [])]
    return {"rows": preview_rows, "fields": fields, "quality": {"score": 88 if preview_rows else 0, "errors": [] if preview_rows else ["文件为空或格式不支持"]}}


def save_mapping(tenant_id: str, shop_id: str, job_id: str, mapping: dict[str, Any]) -> dict[str, Any]:
    execute("UPDATE data_import_jobs SET mapping_json=%s, status='mapping_required', updated_at=NOW() WHERE tenant_id=%s AND shop_id=%s AND id=%s", (json.dumps(mapping, ensure_ascii=False), tenant_id, shop_id, job_id))
    return {"id": job_id, "mappingSaved": True}


def confirm_job(tenant_id: str, shop_id: str, job_id: str, user_id: str = "system") -> dict[str, Any]:
    """确认导入。

    上传 CSV 的确认动作会把预览文件写入正式经营表；示例数据仍由 create_sample_job 直接写入。
    """
    matches = list(IMPORT_DIR.glob(f"{job_id}_*"))
    imported_rows = import_uploaded_csv_data(tenant_id, shop_id, matches[0]) if matches else 0
    execute("UPDATE data_import_jobs SET status='imported', rows_count=%s, quality_score=90, updated_at=NOW() WHERE tenant_id=%s AND shop_id=%s AND id=%s", (imported_rows, tenant_id, shop_id, job_id))
    refresh_result = run_post_import_refresh(tenant_id, shop_id, user_id, job_id)
    return {"id": job_id, "status": "imported", "rows": imported_rows, **refresh_result}


def import_uploaded_csv_data(tenant_id: str, shop_id: str, file_path: Path) -> int:
    """解析上传 CSV/Excel，并写入当前店铺经营表。

    当前支持 orange_shop_import_sample.csv/xlsx 这种宽表格式：一行同时包含订单、商品、库存、流量、
    活动与退款字段。字段缺失时会尽量用默认值兜底，让测试上传流程更宽容。
    """
    rows = _read_upload_rows(file_path)
    if not rows:
        return 0

    prefix = _scope_prefix(tenant_id, shop_id)
    seller_id = f"{prefix}_uploaded_seller"
    now = datetime.now().replace(microsecond=0)
    execute(
        """
        INSERT INTO sellers (tenant_id, shop_id, seller_id, seller_zip_code_prefix, seller_city, seller_state)
        VALUES (%s, %s, %s, 341000, '赣州', '江西')
        ON DUPLICATE KEY UPDATE seller_city=VALUES(seller_city), seller_state=VALUES(seller_state)
        """,
        (tenant_id, shop_id, seller_id),
    )

    customers = {}
    products = {}
    inventories = {}
    orders = []
    order_items = []
    payments = []
    refunds = []
    campaigns = {}
    campaign_stats = {}
    traffic_rows = []

    for index, row in enumerate(rows, start=1):
        source_order_id = _safe_id(row.get("order_id"), f"upload_order_{index:04d}")
        source_customer_id = _safe_id(row.get("customer_id"), f"upload_customer_{index:04d}")
        source_product_id = _safe_id(row.get("product_id") or row.get("sku_code"), f"upload_product_{index:04d}")
        order_id = f"{prefix}_{source_order_id}"[:64]
        customer_id = f"{prefix}_{source_customer_id}"[:64]
        product_id = f"{prefix}_{source_product_id}"[:64]
        order_time = _parse_datetime(row.get("order_time"), now - timedelta(hours=index))
        unit_price = _float(row.get("unit_price"), _float(row.get("pay_amount"), 0))
        quantity = max(1, _int(row.get("quantity"), 1))
        pay_amount = _float(row.get("pay_amount"), unit_price * quantity)
        campaign_name = row.get("campaign_name") or "CSV 上传活动"
        campaign_id = _short_scoped_id(prefix, "campaign", campaign_name, max_len=32)

        customers[customer_id] = (tenant_id, shop_id, customer_id, f"{customer_id}_unique", 330000 + (index % 9000), row.get("customer_city") or "未知城市", "未知")
        products[product_id] = (tenant_id, shop_id, product_id, row.get("product_name") or product_id, row.get("category") or "uploaded", len(row.get("product_name") or product_id))
        inventories[product_id] = (tenant_id, shop_id, product_id, _int(row.get("stock"), 0), _int(row.get("safety_stock"), 0), now)
        orders.append((tenant_id, shop_id, order_id, customer_id, "delivered", order_time, order_time + timedelta(minutes=5), order_time + timedelta(hours=12), order_time + timedelta(days=2), order_time + timedelta(days=4)))
        order_items.append((tenant_id, shop_id, order_id, 1, product_id, seller_id, order_time + timedelta(days=1), pay_amount, 8.0))
        payments.append((tenant_id, shop_id, order_id, 1, "uploaded", 1, pay_amount + 8.0))
        traffic_rows.append((tenant_id, shop_id, order_time.date(), product_id, _int(row.get("visitors"), 0) * 2, _int(row.get("visitors"), 0), max(0, _int(row.get("conversions"), 0) * 2), max(0, _int(row.get("conversions"), 0)), _int(row.get("conversions"), 0)))
        campaigns[campaign_id] = (tenant_id, shop_id, campaign_id, campaign_name, row.get("platform") or "CSV 上传", order_time - timedelta(days=1), order_time + timedelta(days=7), max(1000.0, _float(row.get("ad_spend"), 0) * 10), "active")
        campaign_stats[(campaign_id, product_id)] = (tenant_id, shop_id, campaign_id, product_id, _int(row.get("visitors"), 0) * 3, _int(row.get("visitors"), 0), _int(row.get("conversions"), 0), pay_amount, _float(row.get("ad_spend"), 0))
        if str(row.get("refund_flag") or "").strip().upper() in {"Y", "YES", "TRUE", "1"}:
            refunds.append((tenant_id, shop_id, f"{prefix}_upload_refund_{index:04d}", order_id, product_id, order_time + timedelta(days=2), _float(row.get("refund_amount"), 0), row.get("refund_reason") or "未填写", "approved"))

    execute_many(
        """
        INSERT INTO customers (tenant_id, shop_id, customer_id, customer_unique_id, customer_zip_code_prefix, customer_city, customer_state)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE customer_city=VALUES(customer_city)
        """,
        customers.values(),
    )
    execute_many(
        """
        INSERT INTO products (tenant_id, shop_id, product_id, category_name, category_name_en, product_name_length, product_description_length, product_photos_qty, product_weight_g, product_length_cm, product_height_cm, product_width_cm)
        VALUES (%s, %s, %s, %s, %s, %s, 260, 5, 5000, 35, 22, 28)
        ON DUPLICATE KEY UPDATE category_name=VALUES(category_name), category_name_en=VALUES(category_name_en)
        """,
        products.values(),
    )
    execute_many(
        """
        INSERT INTO inventory (tenant_id, shop_id, product_id, stock, safety_stock, warehouse, updated_at)
        VALUES (%s, %s, %s, %s, %s, 'CSV 上传仓', %s)
        ON DUPLICATE KEY UPDATE stock=VALUES(stock), safety_stock=VALUES(safety_stock), updated_at=VALUES(updated_at)
        """,
        inventories.values(),
    )
    execute_many(
        """
        INSERT INTO orders (tenant_id, shop_id, order_id, customer_id, order_status, order_purchase_timestamp, order_approved_at, order_delivered_carrier_date, order_delivered_customer_date, order_estimated_delivery_date)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE order_status=VALUES(order_status), order_purchase_timestamp=VALUES(order_purchase_timestamp)
        """,
        orders,
    )
    execute_many(
        """
        INSERT INTO order_items (tenant_id, shop_id, order_id, order_item_id, product_id, seller_id, shipping_limit_date, price, freight_value)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE product_id=VALUES(product_id), price=VALUES(price)
        """,
        order_items,
    )
    execute_many(
        """
        INSERT INTO payments (tenant_id, shop_id, order_id, payment_sequential, payment_type, payment_installments, payment_value)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE payment_type=VALUES(payment_type), payment_value=VALUES(payment_value)
        """,
        payments,
    )
    execute_many("INSERT INTO traffic_stats (tenant_id, shop_id, stat_date, product_id, views, visitors, add_to_cart, favorites, conversions) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)", traffic_rows)
    execute_many(
        """
        INSERT INTO campaigns (tenant_id, shop_id, campaign_id, campaign_name, channel, start_time, end_time, budget, status)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE campaign_name=VALUES(campaign_name), status=VALUES(status)
        """,
        campaigns.values(),
    )
    execute_many(
        """
        INSERT INTO campaign_product_stats (tenant_id, shop_id, campaign_id, product_id, impressions, clicks, orders_count, revenue, spend)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE impressions=VALUES(impressions), clicks=VALUES(clicks), orders_count=VALUES(orders_count), revenue=VALUES(revenue), spend=VALUES(spend)
        """,
        campaign_stats.values(),
    )
    execute_many(
        """
        INSERT INTO refunds (tenant_id, shop_id, refund_id, order_id, product_id, refund_time, refund_amount, refund_reason, refund_status)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE refund_status=VALUES(refund_status), refund_amount=VALUES(refund_amount)
        """,
        refunds,
    )
    return len(rows)


def _guess_target(field: str) -> str:
    lowered = field.lower()
    if "order" in lowered:
        return "order_id"
    if "sku" in lowered or "product" in lowered:
        return "product_id"
    if "amount" in lowered or "price" in lowered:
        return "price"
    if "stock" in lowered:
        return "stock"
    return "未识别"


def _read_upload_rows(file_path: Path, limit: int | None = None) -> list[dict[str, str]]:
    """读取 CSV 或 Excel 首个工作表，并统一转成字段字典列表。"""
    suffix = file_path.suffix.lower()
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        from openpyxl import load_workbook

        workbook = load_workbook(file_path, read_only=True, data_only=True)
        worksheet = workbook.active
        iterator = worksheet.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(iterator, [])]
        rows = []
        for index, values in enumerate(iterator):
            if limit is not None and index >= limit:
                break
            item = {header: "" if value is None else str(value) for header, value in zip(headers, values) if header}
            if any(item.values()):
                rows.append(item)
        workbook.close()
        return rows
    with file_path.open("r", encoding="utf-8-sig", newline="") as handle:
        sample = handle.read(4096)
        handle.seek(0)
        delimiter = "\t" if "\t" in sample.splitlines()[0] and "," not in sample.splitlines()[0] else ","
        reader = csv.DictReader(handle, delimiter=delimiter)
        rows = []
        for index, item in enumerate(reader):
            if limit is not None and index >= limit:
                break
            rows.append({key: value or "" for key, value in item.items() if key})
        return rows


def _parse_datetime(value: str | None, fallback: datetime) -> datetime:
    """兼容常见 CSV 时间格式。"""
    if not value:
        return fallback
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M:%S", "%Y/%m/%d"):
        try:
            return datetime.strptime(value.strip(), fmt)
        except ValueError:
            continue
    return fallback


def _float(value: str | int | float | None, fallback: float = 0.0) -> float:
    """把 CSV 单元格转成 float，空值和非法值走默认值。"""
    try:
        if value is None or value == "":
            return fallback
        return float(value)
    except (TypeError, ValueError):
        return fallback


def _int(value: str | int | float | None, fallback: int = 0) -> int:
    """把 CSV 单元格转成 int，支持小数字符串。"""
    return int(_float(value, float(fallback)))


def _safe_id(value: str | None, fallback: str) -> str:
    """把用户上传字段清洗成适合放入 demo 主键的短 ID。"""
    raw = (value or fallback).strip().lower()
    cleaned = "".join(char if char.isalnum() else "_" for char in raw).strip("_")
    return cleaned or fallback


def _short_scoped_id(prefix: str, kind: str, value: str | None, *, max_len: int) -> str:
    """生成兼容旧 schema 短列宽的稳定业务 ID。"""
    digest = hashlib.sha1(f"{prefix}:{kind}:{value or ''}".encode("utf-8")).hexdigest()[:10]
    head = _safe_id(kind, "id")[: max(1, max_len - 11)]
    return f"{head}_{digest}"[:max_len]


def _scope_prefix(tenant_id: str, shop_id: str) -> str:
    """生成稳定且适合放入主键的测试数据前缀。"""
    raw = f"{tenant_id}_{shop_id}".lower()
    return "".join(char if char.isalnum() else "_" for char in raw)[-42:]
